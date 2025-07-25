import sys
import os
import re
import json
import time
import requests
import logging
import openpyxl
import hashlib
import asyncio
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QMessageBox
)
from PyQt5.QtCore import pyqtSignal, QDate
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PyQt5 import uic
from telethon.tl.types import Channel, MessageMediaPhoto, MessageMediaDocument
from telethon.sync import TelegramClient
from telethon.errors import (
    SessionPasswordNeededError, 
    PhoneNumberInvalidError,
    FloodWaitError
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

load_dotenv()
# Константы
# Токены для доступа к API, создаете файл .env в ту же директорию, где лежит файл parse_main.py
# Одноклассники OAuth vk используют... что блять, надо разбираться 
OK_TOKEN = os.getenv("OK_TOKEN")
OK_SECRET_KEY = os.getenv("OK_SECRET_KEY")
HASH_TOKEN = os.getenv("TELEGRAM_API_HASH")
TG_API = os.getenv("TELEGRAM_API_ID")
VK_TOKEN = os.getenv("VK_TOKEN")
VK_VERSION = '5.137'
MAX_POSTS = 100
MAX_WORKERS = 5
REQUEST_DELAY = 0.34
MAX_ATTEMPTS = 3
BACKOFF_FACTOR = 2

# Стили для Excel
HEADER_FILL = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LINK_FONT = Font(color="0000FF", underline="single")

class TelegramParser:
    def __init__(self):
        self.client = None
        self.phone = None
        
    async def auth(self, phone, code=None, password=None):
        try:
            self.client = TelegramClient('session_name', TG_API, HASH_TOKEN)
            await self.client.start(phone=phone, code=code, password=password)
            self.phone = phone
            return True
        except Exception as e:
            logging.error(f"Ошибка авторизации Telegram: {str(e)}")
            return False

    async def get_channel_posts(self, channel_name, start_date, end_date):
        if not self.client:
            raise Exception("Клиент Telegram не инициализирован")

        posts = []
        start_ts = int(time.mktime(start_date.timetuple()))
        end_ts = int(time.mktime(end_date.timetuple()))

        try:
            # Пробуем получить канал по username или ID
            entity = await self.client.get_entity(channel_name)
            
            async for message in self.client.iter_messages(entity):
                if message.date.timestamp() < start_ts:
                    break
                    
                if start_ts <= message.date.timestamp() < end_ts:
                    post = {
                        'text': message.text or "",
                        'date': message.date.strftime('%d.%m.%Y %H:%M'),
                        'views': message.views or 0,
                        'id': message.id,
                        'link': f"https://t.me/{channel_name}/{message.id}"
                    }
                    
                    if message.media:
                        if isinstance(message.media, MessageMediaPhoto):
                            post['media'] = 'photo'
                        elif isinstance(message.media, MessageMediaDocument):
                            post['media'] = 'document'
                    
                    posts.append(post)
                    
        except Exception as e:
            logging.error(f"Ошибка получения постов: {str(e)}")
            
        return posts

#попытка не пытка, мб у них так же как и у вкXD
class OKParser:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        self.access_token = None
        
    def auth(self, access_token):
        """Установка токена доступа"""
        self.access_token = access_token
        
    def get_group_posts(self, group_id, start_date, end_date):
        """Получение постов из группы OK"""
        if not self.access_token:
            raise Exception("Токен доступа OK не установлен")

        posts = []
        start_ts = int(time.mktime(start_date.timetuple()))
        end_ts = int(time.mktime(end_date.timetuple()))
        
        try:
            # Получаем информацию о группе
            params = {
                'application_key': OK_TOKEN,
                'format': 'json',
                'method': 'group.getInfo',
                'gid': group_id,
                'fields': 'name,description',
                'access_token': self.access_token
            }
            
            group_info = self.session.get(
                'https://api.ok.ru/fb.do',
                params=params,
                timeout=10
            ).json()
            
            if 'error' in group_info:
                raise Exception(f"OK API error: {group_info['error']}")
                
            # Получаем посты
            params = {
                'application_key': OK_TOKEN,
                'format': 'json',
                'method': 'group.getFeed',
                'gid': group_id,
                'count': 100,
                'access_token': self.access_token
            }
            
            response = self.session.get(
                'https://api.ok.ru/fb.do',
                params=params,
                timeout=10
            ).json()
            
            if 'error' in response:
                raise Exception(f"OK API error: {response['error']}")
                
            for post in response.get('feed', []):
                if 'created' not in post:
                    continue
                    
                post_time = int(post['created'])
                if post_time < start_ts:
                    continue
                if post_time >= end_ts:
                    continue
                    
                posts.append({
                    'text': post.get('text', ''),
                    'date': datetime.fromtimestamp(post_time).strftime('%d.%m.%Y %H:%M'),
                    'likes': post.get('like_count', 0),
                    'comments': post.get('comments_count', 0),
                    'id': post['id'],
                    'link': f"https://ok.ru/group/{group_id}/topic/{post['id']}"
                })
                
        except Exception as e:
            logging.error(f"Ошибка OK API: {str(e)}")
            
        return posts

class VKAPIError(Exception):
    pass

class VKParser(QMainWindow):
    update_progress = pyqtSignal(int)
    update_status = pyqtSignal(str)
    parsing_finished = pyqtSignal()
    telegram_auth_needed = pyqtSignal()

    def __init__(self):
        """Инициализация парсера"""
        super().__init__()
        self.setup_ui()
        self.setup_connections()
        
        self.thread_pool = ThreadPoolExecutor(max_workers=MAX_WORKERS)
        self.running = False
        self.stop_flag = False
        self.search_texts = []
        self.communities = []
        self.save_folder = os.path.join(os.getcwd(), "результаты_парсинга")
        
        self.telegram_parser = TelegramParser()
        self.ok_parser = OKParser()
        if OK_TOKEN:
            self.ok_parser.auth(OK_TOKEN)
        
        self.last_communities_file = None
        self.last_communities_hash = None
        self.telegram_auth_data = None
        
        if not os.path.exists(self.save_folder):
            os.makedirs(self.save_folder)
        self.load_communities_config()

    def load_communities_config(self):
        """
        Загружает конфигурацию из JSON-файла (путь к последнему файлу сообществ и их хэш)
        Вызывается при инициализации приложения
        """
        config_path = os.path.join(os.getcwd(), "vk_parser_config.json")
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.last_communities_file = config.get('last_communities_file')
                    self.last_communities_hash = config.get('last_communities_hash')
            except Exception as e:
                logging.error(f"Ошибка загрузки конфига: {str(e)}")

    def save_communities_config(self):
        """
        Сохраняет текущую конфигурацию (путь к файлу сообществ и их хэш) в JSON-файл
        Вызывается после успешной загрузки файла с сообществами
        """
        config_path = os.path.join(os.getcwd(), "vk_parser_config.json")
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump({
                    'last_communities_file': self.last_communities_file,
                    'last_communities_hash': self.last_communities_hash
                }, f, ensure_ascii=False, indent=4)
        except Exception as e:
            logging.error(f"Ошибка сохранения конфига: {str(e)}")

    def load_communities_file(self):
        """
        Модифицированная функция загрузки файла с сообществами
        Теперь также сохраняет информацию о загруженном файле и вычисляет хэш списка
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл с сообществами", "", "Excel Files (*.xlsx *.xls)"
        )

        if file_path:
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                self.communities = []
                
                for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                    if not row or len(row) < 2:
                        continue
                    
                    link = str(row[0]).strip() if row[0] else ""
                    name = str(row[1]).strip() if row[1] else ""
                    
                    domain = self.extract_domain_from_link(link)
                    if domain:
                        self.communities.append({
                            "original_link": link,
                            "domain": domain,
                            "name": name
                        })
                
                self.update_status.emit(f"Загружено {len(self.communities)} сообществ")
                
                # Сохраняем информацию о загруженном файле и хэш списка
                self.last_communities_file = file_path
                self.last_communities_hash = self.calculate_communities_hash()
                self.save_communities_config()
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
                logging.error(f"Ошибка загрузки файла: {str(e)}")

    def calculate_communities_hash(self):
        """
        Вычисляет MD5 хэш списка сообществ для сравнения
        Используется для определения, изменился ли список сообществ с прошлого раза
        """
        
        # Сортируем ключи для стабильного хэширования
        communities_str = json.dumps(self.communities, sort_keys=True)
        return hashlib.md5(communities_str.encode('utf-8')).hexdigest()

    def setup_ui(self):
        uic.loadUi("parse_main.ui", self)
        
        self.startDateEdit.setDate(QDate.currentDate())
        self.endDateEdit.setDate(QDate.currentDate())
        self.progressBar.setValue(0)
        self.statusLabel.setText("Готов к работе")

    def setup_connections(self):
        self.parseButton.clicked.connect(self.start_parsing)
        self.selectFolderButton.clicked.connect(self.select_save_folder)
        self.loadCommunitiesButton.clicked.connect(self.load_communities_file)
        
        self.update_progress.connect(self.progressBar.setValue)
        self.update_status.connect(self.statusLabel.setText)
        self.parsing_finished.connect(self.on_parsing_finished)

    def select_save_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        if folder:
            self.save_folder = folder
            self.update_status.emit(f"Папка сохранения: {folder}")

    def load_communities_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл с сообществами", "", "Excel Files (*.xlsx *.xls)"
        )

        if file_path:
            try:
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active
                self.communities = []
                
                for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                    if not row or len(row) < 2:
                        continue
                    
                    link = str(row[0]).strip() if row[0] else ""
                    name = str(row[1]).strip() if row[1] else ""
                    
                    domain = self.extract_domain_from_link(link)
                    if domain:
                        self.communities.append({
                            "original_link": link,
                            "domain": domain,
                            "name": name
                        })
                
                self.update_status.emit(f"Загружено {len(self.communities)} сообществ")
                
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
                logging.error(f"Ошибка загрузки файла: {str(e)}")

    def extract_domain_from_link(self, link):
        if not link:
            return None

        link = str(link).strip().lower()
        
        patterns_vk = [
            r'vk\.com/([a-z0-9_\-\.]+)',
            r'club(\d+)',
            r'public(\d+)',
            r'([a-z0-9_\-\.]+)$'
        ]

        patterns_ok = [
            r'ok\.ru/([a-z0-9_\-\.]+)',
            r'group/([a-z0-9_\-\.]+)',
            r'topic/([a-z0-9_\-\.]+)',
            r'profile/([a-z0-9_\-\.]+)',
            r'video/([a-z0-9_\-\.]+)',
            r'statuses/([a-z0-9_\-\.]+)'
        ]

        patterns_tg = [
            r'telegram.me ([a-z0-9_\-\.]+)',
            r'tg\.me/([a-z0-9_\-\.]+)',
            r't\.me/([a-z0-9_\-\.]+)',
            r'telegram\.org/([a-z0-9_\-\.]+)'
        ]

        for pattern in patterns_vk:
            match = re.search(pattern, link)
            if match:
                domain = match.group(1)
                return f"vk_{domain}"
            
        for pattern in patterns_ok:
            match = re.search(pattern, link)
            if match:
                domain = match.group(1)
                return f"ok_{domain}"
            
        for pattern in patterns_tg:
            match = re.search(pattern, link)
            if match:
                domain = match.group(1)
                return f"tg_{domain}"
        return None
    
    def get_search_texts(self):
        text = self.textEdit.toPlainText().strip()
        return [t.strip() for t in text.split(";") if t.strip()] if text else []

    def get_selected_dates(self):
        start_date = self.startDateEdit.date().toPyDate()
        end_date = self.endDateEdit.date().toPyDate()
        return (start_date, end_date + timedelta(days=1)) if start_date == end_date else (
            (end_date, start_date + timedelta(days=1)) if start_date > end_date else (start_date, end_date)
        )

    def make_vk_request(self, method, params):
        for attempt in range(MAX_ATTEMPTS):
            try:
                time.sleep(REQUEST_DELAY * (BACKOFF_FACTOR ** attempt))
                params.update({'access_token': VK_TOKEN, 'v': VK_VERSION})
                response = requests.get(f'https://api.vk.com/method/{method}', params=params, timeout=10)
                data = response.json()
                
                if 'error' in data:
                    error = data['error']
                    if error['error_code'] == 6:  # Too many requests
                        time.sleep(1 + attempt)
                        continue
                    elif error['error_code'] in [5, 15, 18, 100]:
                        raise VKAPIError(f"VK API error {error['error_code']}: {error['error_msg']}")
                    continue
                    
                return data
            
            except requests.exceptions.RequestException as e:
                logging.error(f"Сетевая ошибка: {str(e)}")
                continue
        
        raise VKAPIError(f"Не удалось выполнить запрос после {MAX_ATTEMPTS} попыток")

    def get_group_posts(self, domain, start_date, end_date):
        if not domain.startswith('vk_'):
            logging.error(f"Некорректный домен для VK: {domain}")
            return []
        
        vk_domain = domain[3:]
        posts = []
        offset = 0
        start_ts = int(time.mktime(start_date.timetuple()))
        end_ts = int(time.mktime(end_date.timetuple()))
        
        while offset < MAX_POSTS and not self.stop_flag:
            try:
                data = self.make_vk_request('wall.get', {
                    'domain': vk_domain,  # Используем домен без префикса
                    'count': min(100, MAX_POSTS - offset),
                    'offset': offset,
                    'filter': 'owner'
                })
                
                if not data or 'response' not in data:
                    break
                    
                items = data['response'].get('items', [])
                posts.extend(post for post in items if start_ts <= post['date'] < end_ts)
                offset += len(items)
                
                if len(items) < 100:
                    break
                    
            except Exception as e:
                logging.error(f"Ошибка при получении постов: {str(e)} {domain}")
                break
                
        return posts

    def search_text_in_posts(self, posts, search_texts):
        results = []
        search_texts = [t.lower() for t in search_texts]
        
        for post in posts:
            if self.stop_flag:
                break
                
            text = post.get('text', '').lower()
            found_words = [word for word in search_texts if word in text]
            
            if found_words:
                results.append({
                    'post_id': post['id'],
                    'owner_id': post['owner_id'],
                    'text': post.get('text', ''),
                    'date': datetime.fromtimestamp(post['date']).strftime('%d.%m.%Y %H:%M'),
                    'views': post.get('views', {}).get('count', 0),
                    'likes': post.get('likes', {}).get('count', 0),
                    'reposts': post.get('reposts', {}).get('count', 0),
                    'link': f"https://vk.com/wall{post['owner_id']}_{post['id']}",
                    'found_words': ', '.join(found_words)  # слова по которым найден текст
                })
                
        return results

    async def authenticate_telegram(self, phone, code=None, password=None):
        """Асинхронная аутентификация в Telegram"""
        try:
            return await self.telegram_parser.auth(phone, code, password)
        except Exception as e:
            logging.error(f"Ошибка аутентификации Telegram: {str(e)}")
            return False

    def get_telegram_posts(self, domain, start_date, end_date):
        """Получение постов из Telegram канала"""
        if not domain.startswith('tg_'):
            logging.error(f"Некорректный домен для Telegram: {domain}")
            return []
            
        channel_name = domain[3:]
        
        if not self.telegram_parser.client:
            if not self.show_telegram_auth_dialog():
                return []
        
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            posts = loop.run_until_complete(
                self.telegram_parser.get_channel_posts(channel_name, start_date, end_date)
            )
            return posts
        except Exception as e:
            logging.error(f"Ошибка получения постов Telegram: {str(e)}")
            return []
        finally:
            loop.close()

    def show_telegram_auth_dialog(self):
        """Показывает диалог аутентификации Telegram и возвращает True если успешно"""
        auth_dialog = uic.loadUi("telegram_auth.ui")
        auth_dialog.setWindowTitle("Авторизация в Telegram")
        result = False
        
        def handle_submit():
            nonlocal result
            phone = auth_dialog.phoneEdit.text().strip()
            code = auth_dialog.codeEdit.text().strip() if not auth_dialog.codeEdit.isHidden() else None
            password = auth_dialog.passwordEdit.text().strip() if not auth_dialog.passwordEdit.isHidden() else None
            
            if not phone:
                QMessageBox.warning(auth_dialog, "Ошибка", "Введите номер телефона")
                return
                
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                auth_result = loop.run_until_complete(
                    self.telegram_parser.auth(phone, code, password)
                )
                
                if auth_result:
                    result = True
                    auth_dialog.accept()
                else:
                    if not auth_dialog.codeEdit.isVisible():
                        auth_dialog.codeLabel.setHidden(False)
                        auth_dialog.codeEdit.setHidden(False)
                        auth_dialog.phoneEdit.setEnabled(False)
                        QMessageBox.information(auth_dialog, "Информация", "Введите код из Telegram")
                    else:
                        QMessageBox.warning(auth_dialog, "Ошибка", "Не удалось авторизоваться")
            except SessionPasswordNeededError:
                auth_dialog.passwordLabel.setHidden(False)
                auth_dialog.passwordEdit.setHidden(False)
                QMessageBox.information(auth_dialog, "Информация", "Введите пароль 2FA")
            except Exception as e:
                QMessageBox.warning(auth_dialog, "Ошибка", f"Ошибка авторизации: {str(e)}")
            finally:
                loop.close()
        
        auth_dialog.submitButton.clicked.connect(handle_submit)
        auth_dialog.cancelButton.clicked.connect(auth_dialog.reject)
        
        auth_dialog.exec_()
        return result

    def get_ok_posts(self, domain, start_date, end_date):
        """Получение постов из OK группы"""
        if not domain.startswith('ok_'):
            logging.error(f"Некорректный домен для OK: {domain}")
            return []
            
        group_id = domain[3:]
        
        try:
            # Пытаемся преобразовать group_id в число (если это числовой ID)
            try:
                group_id = int(group_id)
            except ValueError:
                pass  # Оставляем как строку, если это не число
                
            return self.ok_parser.get_group_posts(group_id, start_date, end_date)
        except Exception as e:
            logging.error(f"Ошибка получения постов OK: {str(e)}")
            return []

    def process_community(self, community, search_texts, start_date, end_date):
        if self.stop_flag:
            return None
        try:
            posts = []
            if community['domain'].startswith('vk_'):
                posts = self.get_group_posts(community['domain'], start_date, end_date)
            elif community['domain'].startswith('tg_'):
                posts = self.get_telegram_posts(community['domain'], start_date, end_date)
            elif community['domain'].startswith('ok_'):
                posts = self.get_ok_posts(community['domain'], start_date, end_date)
            else:
                return None
            
            if not posts:
                return None
                
            results = self.search_text_in_posts(posts, search_texts)
            return {'community': community, 'results': results} if results else None
            
        except Exception as e:
            logging.error(f"Ошибка обработки сообщества {community['domain']}: {str(e)}")
            return None

    def show_telegram_auth_dialog(self):
        """Показывает диалог аутентификации Telegram"""
        auth_dialog = uic.loadUi("telegram_auth.ui")
        auth_dialog.setWindowTitle("Авторизация в Telegram")
        
        def handle_submit():
            phone = auth_dialog.phoneEdit.text().strip()
            code = auth_dialog.codeEdit.text().strip() if not auth_dialog.codeEdit.isHidden() else None
            password = auth_dialog.passwordEdit.text().strip() if not auth_dialog.passwordEdit.isHidden() else None
            
            if not phone:
                QMessageBox.warning(auth_dialog, "Ошибка", "Введите номер телефона")
                return
            
            # Сохраняем данные для последующих попыток
            self.telegram_auth_data = {
                'phone': phone,
                'code': code,
                'password': password
            }
            
            # Пытаемся авторизоваться
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                auth_result = loop.run_until_complete(
                    self.authenticate_telegram(phone, code, password)
                )
                
                if auth_result:
                    auth_dialog.accept()
                else:
                    # Показываем поле для кода, если требуется
                    if not auth_dialog.codeEdit.isVisible():
                        auth_dialog.codeLabel.setHidden(False)
                        auth_dialog.codeEdit.setHidden(False)
                        auth_dialog.phoneEdit.setEnabled(False)
                        QMessageBox.information(auth_dialog, "Информация", "Введите код из Telegram")
                    else:
                        QMessageBox.warning(auth_dialog, "Ошибка", "Не удалось авторизоваться")
            except SessionPasswordNeededError:
                auth_dialog.passwordLabel.setHidden(False)
                auth_dialog.passwordEdit.setHidden(False)
                QMessageBox.information(auth_dialog, "Информация", "Введите пароль 2FA")
            except Exception as e:
                QMessageBox.warning(auth_dialog, "Ошибка", f"Ошибка авторизации: {str(e)}")
            finally:
                loop.close()
        
        auth_dialog.submitButton.clicked.connect(handle_submit)
        auth_dialog.cancelButton.clicked.connect(auth_dialog.reject)
        
        return auth_dialog.exec_()

    def setup_connections(self):
        self.parseButton.clicked.connect(self.start_parsing)
        self.selectFolderButton.clicked.connect(self.select_save_folder)
        self.loadCommunitiesButton.clicked.connect(self.load_communities_file)
        
        self.update_progress.connect(self.progressBar.setValue)
        self.update_status.connect(self.statusLabel.setText)
        self.parsing_finished.connect(self.on_parsing_finished)
        self.telegram_auth_needed.connect(self.show_telegram_auth_dialog)

    def process_community(self, community, search_texts, start_date, end_date):
        if self.stop_flag:
            return None
        try:
            if community['domain'].startswith('vk_'):
                posts = self.get_group_posts(community['domain'], start_date, end_date)
            elif community['domain'].startswith('tg_'):
                posts = self.get_telegram_posts(community['domain'], start_date, end_date)
            elif community['domain'].startswith('ok_'):
                posts = self.get_ok_posts(community['domain'], start_date, end_date)
            else:
                return None
            
            if not posts:
                return None
                
            results = self.search_text_in_posts(posts, search_texts)
            return {'community': community, 'results': results} if results else None
            
        except Exception as e:
            logging.error(f"Ошибка обработки сообщества {community['domain']}: {str(e)}")
            return None

    def create_empty_communities_sheet(self, wb, empty_communities):
        if not empty_communities:
            return
            
        ws = wb.create_sheet(title="Не найдено")
        headers = ["Ссылка", "Название", "Причина"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
        
        for row_num, comm in enumerate(empty_communities, 2):
            ws.cell(row=row_num, column=1, value=comm['original_link']).font = LINK_FONT
            ws.cell(row=row_num, column=2, value=comm['name'])
            ws.cell(row=row_num, column=3, value=comm.get('reason', 'Нет совпадений'))
        
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    def create_report(self, data, empty_communities):
        """
        функция создания отчета
        сохраняет все результаты в один файл с разными листами
        Логика работы:
        1. Всегда использует файл 'результаты_парсинга.xlsx' в папке сохранения
        2. Если файл существует и список сообществ не изменился:
        - Для текущей даты создает/перезаписывает лист с результатами
        - Для пустых сообществ создает лист с датой в названии
        3. Если список сообществ изменился - создает новый файл
        P.S. странно немного но окэ
        """
        if not data and not empty_communities:
            return None
            
        try:
            current_date = datetime.now().strftime("%d.%m.%Y")
            report_name = "результаты_парсинга.xlsx"
            filepath = os.path.join(self.save_folder, report_name)
            
            same_communities = False
            if os.path.exists(filepath) and self.last_communities_hash:
                current_hash = self.calculate_communities_hash()
                if current_hash == self.last_communities_hash:
                    same_communities = True
            
            if same_communities and os.path.exists(filepath):
                wb = openpyxl.load_workbook(filepath)
                if current_date in wb.sheetnames:
                    del wb[current_date]
                ws = wb.create_sheet(title=current_date)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = current_date
            
            if data:
                headers = [
                    'Ссылка на сообщество', 'Название', 'Ссылка на пост',
                    'Текст поста', 'Найденные слова', 'Дата', 'Просмотры', 'Лайки', 'Репосты'
                ]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = HEADER_ALIGN
                
                row_num = 2
                for item in data:
                    comm = item['community']
                    for res in item['results']:
                        ws.cell(row=row_num, column=1, value=comm['original_link']).font = LINK_FONT
                        ws.cell(row=row_num, column=2, value=comm['name'])
                        ws.cell(row=row_num, column=3, value=res['link']).font = LINK_FONT
                        text_cell = ws.cell(row=row_num, column=4, value=res['text'])
                        text_cell.alignment = Alignment(wrap_text=True)
                        ws.cell(row=row_num, column=5, value=res['found_words'])  # Новая колонка с найденными словами
                        ws.cell(row=row_num, column=6, value=res['date'])
                        ws.cell(row=row_num, column=7, value=res['views'])
                        ws.cell(row=row_num, column=8, value=res['likes'])
                        ws.cell(row=row_num, column=9, value=res['reposts'])
                        row_num += 1
                
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            empty_sheet_name = f"Не найдено ({current_date})"
            if empty_sheet_name in wb.sheetnames:
                del wb[empty_sheet_name]
                
            if empty_communities:
                ws_empty = wb.create_sheet(title=empty_sheet_name)
                headers = ["Ссылка", "Название", "Причина"]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws_empty.cell(row=1, column=col_num, value=header)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = HEADER_ALIGN
                
                for row_num, comm in enumerate(empty_communities, 2):
                    ws_empty.cell(row=row_num, column=1, value=comm['original_link']).font = LINK_FONT
                    ws_empty.cell(row=row_num, column=2, value=comm['name'])
                    ws_empty.cell(row=row_num, column=3, value=comm.get('reason', 'Нет совпадений'))
                

                for column in ws_empty.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws_empty.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            logging.error(f"Ошибка создания отчета: {str(e)}")
            return None

    def start_parsing(self):
        if self.running:
            return
            
        self.search_texts = self.get_search_texts()
        if not self.search_texts:
            QMessageBox.warning(self, "Ошибка", "Введите текст для поиска")
            return
            
        if not self.communities:
            QMessageBox.warning(self, "Ошибка", "Загрузите список сообществ")
            return
            
        self.running = True
        self.stop_flag = False
        self.parseButton.setEnabled(False)
        self.progressBar.setValue(0)
        
        start_date, end_date = self.get_selected_dates()
        self.update_status.emit(f"Парсинг с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}")
        
        self.thread_pool.submit(self.run_parsing, start_date, end_date)

    def run_parsing(self, start_date, end_date):
        try:
            vk_communities = [comm for comm in self.communities if comm['domain'].startswith('vk_')]
            total = len(vk_communities)
            results = []
            empty_communities = []
            
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = {
                    executor.submit(
                        self.process_community,
                        comm,
                        self.search_texts,
                        start_date,
                        end_date
                    ): comm for comm in vk_communities
                }
                
                for i, future in enumerate(as_completed(futures), 1):
                    if self.stop_flag:
                        break
                        
                    result = future.result()
                    if result:
                        results.append(result)
                    else:
                        empty_communities.append({
                            'original_link': futures[future]['original_link'],
                            'name': futures[future]['name'],
                            'reason': 'Нет совпадений'
                        })
                    
                    progress = int((i / total) * 100)
                    self.update_progress.emit(progress)
                    self.update_status.emit(
                        f"Обработано {i}/{total}. Найдено: {len(results)}, Пустых: {len(empty_communities)}"
                    )
            
            if not self.stop_flag:
                report_path = self.create_report(results, empty_communities)
                if report_path:
                    self.update_status.emit(f"Отчет сохранен: {report_path}")
                    if sys.platform == "win32":
                        os.startfile(self.save_folder)
            
        except Exception as e:
            self.update_status.emit(f"Ошибка: {str(e)}")
            logging.error(f"Ошибка парсинга: {str(e)}")
        finally:
            self.parsing_finished.emit()

    def on_parsing_finished(self):
        self.running = False
        self.parseButton.setEnabled(True)
        self.update_progress.emit(100)
        if not self.stop_flag:
            self.update_status.emit("Парсинг завершен")

    def stop_parsing(self):
        self.stop_flag = True
        self.update_status.emit("Остановка...")

    def closeEvent(self, event):
        if self.running:
            reply = QMessageBox.question(
                self, 'Подтверждение',
                'Парсинг выполняется. Закрыть программу?',
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                self.stop_parsing()
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    parser = VKParser()
    parser.show()
    sys.exit(app.exec_())


###########################
##  WITH LOVE FROM Vi_Ti ##
##          2025         ##
###########################
